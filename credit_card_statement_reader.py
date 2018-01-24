import openpyxl
import PyPDF2
import re
from datetime import datetime


def export_page_data(file_name):

    pdfFileObj = open(file_name, "rb")
    pdfreader = PyPDF2.PdfFileReader(pdfFileObj)
    number_of_pages = pdfreader.numPages
    usable_pages = []
    page_number = 0

    while page_number < number_of_pages:
        pageobj = pdfreader.getPage(page_number)
        page = pageobj.extractText()
        usable_pages.append(page)
        page_number += 1
    for item in usable_pages:
        if "Description$ Amount" not in item:
            del usable_pages[usable_pages.index(item)]

    return usable_pages


def extract_usable_text(extracted_pages):

    edited_text = []

    for item in extracted_pages:
        text = item.replace(",", "")
        payment_splitter = re.findall("Y*o*u*-M*o*b*i*l*e*-\d+.\d\d", text)

        if "Description$ Amount" in text:
            split_text = text.split("Description$ Amount")
            text = str(split_text[1])
        if "Thank You-Mobile-" in text:
            split_text = text.split(payment_splitter[0])
            text = str(split_text[1])
        if "Total fees charged in" in text:
            split_text = text.split("Total fees charged in")
            text = str(split_text[0])
        if "CARDMEMBER SERVICEPO" in text:
            split_text = text.split("CARDMEMBER SERVICEPO")
            text = str(split_text[0])
        edited_text.append(text)

    combined_text = "".join(edited_text)

    return combined_text


def extract_descriptions(text, dates, dollar_amounts):

    values_replaced = text

    for item in dates:
        if item in text:
            values_replaced = values_replaced.replace(item, "")
    for item in dollar_amounts:
        if item in text:
            values_replaced = values_replaced.replace(item, "XxXxX")
    values_replaced = values_replaced.split("XxXxX")
    values_replaced = values_replaced[:len(values_replaced)-1]

    return values_replaced


def extract_dollar_amounts(text):

    string_dollar_amounts = re.findall("\d+\.\d\d", text)
    float_dollar_amounts = []

    for item in string_dollar_amounts:
        float_dollar_amounts.append(float(item))

    return string_dollar_amounts, float_dollar_amounts


def extract_dates(text, year):

    improper_dates = re.findall("\D\d/\d\d|\D\d\d/\d\d|\d+/\d\d", text)
    shortened_dates = []
    shortened_dates.append(improper_dates[0])
    dates_with_years = []
    date_year_string = '/' + year[-2] + year[-1]

    for item in improper_dates:
        if re.search('[a-zA-Z]', item):
            shortened_dates.append(item[1:])
        else:
            shortened_dates.append(item[2:])

    del shortened_dates[1]

    for item in shortened_dates:
        dates_with_years.append(item + date_year_string)

    return dates_with_years, shortened_dates


def export_to_excel(dates, descriptions, dollar_amounts):

    wb = openpyxl.load_workbook('/home/cesare/Documents/exported_bank_statement.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')

    formatter_string = "%m/%d/%y"

    row = 2
    for item in dates:
        datetime_object = datetime.strptime(item, formatter_string)
        sheet.cell(row=row, column=1, value=datetime_object.date())
        row += 1
    row = 2
    for item in descriptions:
        sheet.cell(row=row, column=2, value=item)
        row += 1
    row = 2
    for item in dollar_amounts:
        sheet.cell(row=row, column=3, value=item).number_format = "$#,##0.00"
        row += 1

    wb.save('/home/cesare/Documents/exported_bank_statement.xlsx')


file_input = input("What is the file name?")
file_name = "/home/cesare/Documents/{}.pdf".format(file_input)
all_pages = export_page_data(file_name)
usable_text = extract_usable_text(all_pages)
usable_dates = extract_dates(usable_text, file_input)
usable_dollar_amounts = extract_dollar_amounts(usable_text)
usable_descriptions = extract_descriptions(usable_text, usable_dates[1], usable_dollar_amounts[0])

export_to_excel(usable_dates[0], usable_descriptions, usable_dollar_amounts[1])
